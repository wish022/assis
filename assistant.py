import os
import json
import tkinter as tk
from tkinter import ttk, scrolledtext, messagebox, simpledialog, filedialog, Toplevel, Label, Button, Frame, Entry
import requests
import threading
import imaplib
import email
from email.header import decode_header
from bs4 import BeautifulSoup
from PIL import Image, ImageFilter, ImageOps
import openpyxl
from difflib import SequenceMatcher

# ---------- 本地 Ollama 配置 ----------
OLLAMA_URL = "http://localhost:11434/api/generate"
MODEL = "qwen2.5:7b"   # 可改为 "qwen2.5:1.5b"

def call_llm(prompt, timeout=300):
    try:
        r = requests.post(OLLAMA_URL, json={"model": MODEL, "prompt": prompt, "stream": False}, timeout=timeout)
        r.raise_for_status()
        return r.json()["response"]
    except Exception as e:
        return f"调用失败: {e}"

# ---------- 通用工具 ----------
def read_file(filename):
    if not os.path.exists(filename):
        return None
    with open(filename, "r", encoding="utf-8") as f:
        return f.read()

def load_config():
    if os.path.exists("config.json"):
        with open("config.json", "r", encoding="utf-8") as f:
            return json.load(f)
    return {"email_accounts": [], "fetch_count": 5, "pushplus_token": ""}

def save_config(config):
    with open("config.json", "w", encoding="utf-8") as f:
        json.dump(config, f, indent=4, ensure_ascii=False)

# ---------- 微信推送 ----------
def send_to_wechat(content, token):
    if not token:
        return "未配置 pushplus token"
    url = "http://www.pushplus.plus/send"
    data = {"token": token, "title": "AI分析结果", "content": content, "template": "markdown"}
    try:
        r = requests.post(url, json=data, timeout=10)
        if r.json().get("code") == 200:
            return "已发送到微信"
        else:
            return f"发送失败: {r.json().get('msg')}"
    except Exception as e:
        return f"发送异常: {e}"

# ---------- 邮件读取（IMAP）----------
def fetch_emails_from_account(account_cfg, count=5):
    try:
        mail = imaplib.IMAP4_SSL(account_cfg["server"])
        mail.login(account_cfg["email"], account_cfg["password"])
        mail.select(account_cfg.get("folder", "INBOX"))
        status, messages = mail.search(None, "UNSEEN")
        if status != "OK":
            return False, "无法搜索邮件"
        msg_ids = messages[0].split()
        if not msg_ids:
            return True, []
        if len(msg_ids) > count:
            msg_ids = msg_ids[-count:]
        emails = []
        for msg_id in reversed(msg_ids):
            status, msg_data = mail.fetch(msg_id, "(RFC822)")
            if status != "OK":
                continue
            raw_email = msg_data[0][1]
            msg = email.message_from_bytes(raw_email)
            from_ = decode_header(msg.get("From", ""))[0][0]
            if isinstance(from_, bytes):
                from_ = from_.decode("utf-8", errors="ignore")
            subject = decode_header(msg.get("Subject", ""))[0][0]
            if isinstance(subject, bytes):
                subject = subject.decode("utf-8", errors="ignore")
            body = ""
            if msg.is_multipart():
                for part in msg.walk():
                    ct = part.get_content_type()
                    if ct == "text/plain":
                        body = part.get_payload(decode=True).decode("utf-8", errors="ignore")
                        break
                    elif ct == "text/html":
                        html = part.get_payload(decode=True).decode("utf-8", errors="ignore")
                        soup = BeautifulSoup(html, "html.parser")
                        body = soup.get_text()
            else:
                ct = msg.get_content_type()
                if ct == "text/plain":
                    body = msg.get_payload(decode=True).decode("utf-8", errors="ignore")
                elif ct == "text/html":
                    html = msg.get_payload(decode=True).decode("utf-8", errors="ignore")
                    soup = BeautifulSoup(html, "html.parser")
                    body = soup.get_text()
            emails.append((from_, subject, body.strip()[:2000]))
        mail.close()
        mail.logout()
        return True, emails
    except Exception as e:
        return False, f"连接失败: {str(e)}"

# ---------- 功能1：会议纪要 ----------
def summarize_meeting():
    text = read_file("会议纪要.txt")
    if not text:
        return "找不到会议纪要.txt"
    prompt = f"整理以下会议记录：1.核心摘要 2.待办事项 3.遗留问题\n\n{text}"
    return call_llm(prompt)

# ---------- 功能2：邮件提取 ----------
def extract_email(selected_account_cfg):
    success, result = fetch_emails_from_account(selected_account_cfg)
    if not success:
        return result
    if not result:
        return "📭 没有未读邮件"
    combined = "\n\n".join([f"发件人: {f}\n主题: {s}\n正文: {b}" for f, s, b in result])
    prompt = f"请提取以下邮件中的重点信息（逐条列出）：发件人意图、需要我做什么、截止时间。\n\n{combined}"
    return call_llm(prompt)

# ---------- 功能3：周报生成 ----------
def generate_weekly():
    text = read_file("本周记录.txt")
    if not text:
        return "找不到本周记录.txt"
    prompt = f"根据以下工作记录生成周报（本周工作、问题、下周计划）：\n\n{text}"
    return call_llm(prompt)

# ---------- 功能4：Excel客户分析（选择性波峰焊）----------
def analyze_excel_custom(file_path, custom_prompt=""):
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return "Excel为空"
        content_lines = []
        for i, row in enumerate(rows[:200]):
            row_text = "\t".join([str(cell) if cell else "" for cell in row])
            content_lines.append(row_text)
        table_text = "\n".join(content_lines)
        if not custom_prompt.strip():
            goal = "筛选出可能需要使用选择性波峰焊设备的企业。选择性波峰焊常用于：电子制造服务(EMS)、汽车电子(ECU/车灯/传感器)、电源模块/充电桩、通信设备(基站/路由器)、工业控制(PLC)、家电控制板、LED驱动、医疗电子。"
        else:
            goal = custom_prompt.strip()
        prompt = f"请分析以下参会公司名单，{goal}\n输出Markdown表格：公司名称 | 电话 | 邮箱 | 地址 | 推荐理由\n\n名单内容：\n{table_text}"
        return call_llm(prompt)
    except Exception as e:
        return f"读取Excel失败: {e}"

# ---------- 功能5：产品分析（带轻量记忆）----------
class ProductMemory:
    def __init__(self, file="product_memory.json"):
        self.file = file
        self.data = json.load(open(file, encoding='utf-8')) if os.path.exists(file) else []
    def save(self):
        with open(self.file, 'w', encoding='utf-8') as f:
            json.dump(self.data, f, indent=2, ensure_ascii=False)
    def add(self, title, content):
        self.data.append({"title": title, "content": content, "text": title+content})
        self.save()
    def search(self, query, top_k=3):
        if not self.data:
            return []
        scored = [(SequenceMatcher(None, query.lower(), d['text'].lower()).ratio(), d) for d in self.data]
        scored.sort(reverse=True)
        return [d for _, d in scored[:top_k]]

product_mem = ProductMemory()

def analyze_product(product_desc):
    related = product_mem.search(product_desc)
    ctx = "\n".join([r['content'][:200] for r in related]) if related else ""
    prompt = f"""你是一位资深的产品顾问。请分析以下产品描述，并给出专业评价。
历史参考：{ctx}
产品描述：{product_desc}
输出格式：
1. 产品核心特点
2. 潜在应用场景
3. 优势与不足
4. 改进建议
5. 综合评分（1-10分）
回答要简洁、有洞察。"""
    result = call_llm(prompt)
    product_mem.add(product_desc[:30], result)
    return result

# ---------- 功能6：合同分析（乙方立场，分块）----------
def read_docx(path):
    from docx import Document
    doc = Document(path)
    return "\n".join(p.text for p in doc.paragraphs)

def chunk_text(text, size=3000):
    paras = text.split('\n')
    chunks = []
    cur = []
    cur_len = 0
    for p in paras:
        l = len(p)
        if cur_len + l > size and cur:
            chunks.append("\n".join(cur))
            cur = [p]
            cur_len = l
        else:
            cur.append(p)
            cur_len += l
    if cur:
        chunks.append("\n".join(cur))
    return chunks

def analyze_contract(file_path):
    ext = os.path.splitext(file_path)[1].lower()
    if ext == '.docx':
        try:
            from docx import Document
        except ImportError:
            return "请安装 python-docx: pip install python-docx"
        full_text = read_docx(file_path)
    elif ext in ('.xlsx', '.xls'):
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb.active
        rows = list(sheet.iter_rows(values_only=True))
        full_text = "\n".join(["\t".join(str(c) if c else "" for c in row) for row in rows[:1000]])
    else:
        return "仅支持 .docx 或 .xlsx"
    if not full_text.strip():
        return "文件为空"
    chunks = chunk_text(full_text, size=3000)
    results = []
    for i, ch in enumerate(chunks, 1):
        prompt = f"""你是合同审核专家，代表乙方。分析以下片段（{i}/{len(chunks)}），找出对乙方不利的条款，提取金额。输出表格：条款 | 风险等级 | 涉及金额 | 风险说明 | 修改建议。
片段：{ch}"""
        res = call_llm(prompt, timeout=180)
        results.append(f"## 片段 {i}\n{res}")
# ---------- 邮箱设置、账号管理界面 ----------
def open_email_settings(parent, app_instance):
    win = Toplevel(parent)
    win.title("邮箱账号设置")
    win.geometry("500x550")
    win.configure(bg='#2d2d2d')
    win.transient(parent)
    win.grab_set()
    config = load_config()
    accounts = config.get("email_accounts", [])
    fetch_count = config.get("fetch_count", 5)
    pushplus_token = config.get("pushplus_token", "")
    tk.Label(win, text="已配置的邮箱账号:", bg='#2d2d2d', fg='white', font=("微软雅黑", 10)).pack(anchor='w', padx=10, pady=(10,0))
    listbox_frame = tk.Frame(win, bg='#2d2d2d')
    listbox_frame.pack(fill='both', expand=True, padx=10, pady=5)
    listbox = tk.Listbox(listbox_frame, bg='#3c3c3c', fg='white', selectbackground='#1e1e1e')
    listbox.pack(side='left', fill='both', expand=True)
    scrollbar = tk.Scrollbar(listbox_frame, orient='vertical', command=listbox.yview)
    scrollbar.pack(side='right', fill='y')
    listbox.config(yscrollcommand=scrollbar.set)
    for acc in accounts:
        listbox.insert(tk.END, f"{acc.get('name', '未命名')} - {acc.get('email', '')}")
    selected_idx = None
    def on_select(evt):
        nonlocal selected_idx
        sel = listbox.curselection()
        if sel:
            selected_idx = sel[0]
    listbox.bind('<<ListboxSelect>>', on_select)
    btn_frame = tk.Frame(win, bg='#2d2d2d')
    btn_frame.pack(pady=5)
    def add_account():
        edit_window(win, None, app_instance)
    def edit_account():
        if selected_idx is None:
            messagebox.showinfo("提示", "请先选择一个账号")
            return
        edit_window(win, selected_idx, app_instance)
    def delete_account():
        nonlocal selected_idx
        if selected_idx is None:
            messagebox.showinfo("提示", "请先选择一个账号")
            return
        if messagebox.askyesno("确认删除", f"删除账号 {accounts[selected_idx].get('name', '')} 吗？"):
            accounts.pop(selected_idx)
            config["email_accounts"] = accounts
            save_config(config)
            app_instance.refresh_accounts()
            win.destroy()
            open_email_settings(parent, app_instance)
    tk.Button(btn_frame, text="新增账号", command=add_account, bg='#4CAF50', fg='white', relief=tk.FLAT).pack(side='left', padx=5)
    tk.Button(btn_frame, text="编辑选中", command=edit_account, bg='#2196F3', fg='white', relief=tk.FLAT).pack(side='left', padx=5)
    tk.Button(btn_frame, text="删除选中", command=delete_account, bg='#f44336', fg='white', relief=tk.FLAT).pack(side='left', padx=5)

    token_frame = tk.Frame(win, bg='#2d2d2d')
    token_frame.pack(pady=10, fill='x', padx=10)
    tk.Label(token_frame, text="pushplus token (用于微信推送):", bg='#2d2d2d', fg='white').pack(anchor='w')
    token_entry = tk.Entry(token_frame, width=40, bg='#3c3c3c', fg='white', insertbackground='white')
    token_entry.insert(0, pushplus_token)
    token_entry.pack(fill='x', pady=5)
    def save_token():
        new_token = token_entry.get().strip()
        config["pushplus_token"] = new_token
        save_config(config)
        app_instance.pushplus_token = new_token
        messagebox.showinfo("成功", "token已保存")
    tk.Button(token_frame, text="保存 token", command=save_token, bg='#FF9800', fg='white', relief=tk.FLAT).pack(pady=5)

    count_frame = tk.Frame(win, bg='#2d2d2d')
    count_frame.pack(pady=10, fill='x', padx=10)
    tk.Label(count_frame, text="每次获取未读邮件数量:", bg='#2d2d2d', fg='white').pack(side='left')
    count_var = tk.StringVar(value=str(fetch_count))
    count_entry = tk.Entry(count_frame, textvariable=count_var, width=5)
    count_entry.pack(side='left', padx=5)
    def save_count():
        try:
            new_count = int(count_var.get())
            config["fetch_count"] = new_count
            save_config(config)
            app_instance.fetch_count = new_count
            messagebox.showinfo("成功", "已保存")
        except:
            messagebox.showerror("错误", "请输入数字")
    tk.Button(count_frame, text="保存数量", command=save_count, bg='#FF9800', fg='white', relief=tk.FLAT).pack(side='left', padx=5)

    tk.Button(win, text="关闭", command=win.destroy, bg='#3c3c3c', fg='white', relief=tk.FLAT).pack(pady=10)

def edit_window(parent, idx, app_instance):
    config = load_config()
    accounts = config.get("email_accounts", [])
    if idx is None:
        title = "新增邮箱账号"
        account = {"name": "", "server": "imap.qq.com", "email": "", "password": "", "folder": "INBOX"}
    else:
        title = "编辑邮箱账号"
        account = accounts[idx].copy()
    win = Toplevel(parent)
    win.title(title)
    win.geometry("400x350")
    win.configure(bg='#2d2d2d')
    win.transient(parent)
    win.grab_set()
    fields = [("账号名称 (例如: 工作邮箱)", "name"), ("IMAP 服务器地址", "server"), ("邮箱地址", "email"), ("授权码 (不是登录密码)", "password")]
    entries = {}
    for i, (label, key) in enumerate(fields):
        tk.Label(win, text=label, bg='#2d2d2d', fg='white', anchor='w').pack(fill='x', padx=10, pady=(5,0))
        entry = tk.Entry(win, width=50, bg='#3c3c3c', fg='white', insertbackground='white')
        entry.insert(0, account.get(key, ""))
        entry.pack(padx=10, pady=(0,5), fill='x')
        entries[key] = entry
    def save():
        new_account = {}
        for key in ["name", "server", "email", "password"]:
            val = entries[key].get().strip()
            if not val:
                messagebox.showerror("错误", f"{key} 不能为空")
                return
            new_account[key] = val
        new_account["folder"] = "INBOX"
        if idx is None:
            accounts.append(new_account)
        else:
            accounts[idx] = new_account
        config["email_accounts"] = accounts
        save_config(config)
        app_instance.refresh_accounts()
        win.destroy()
        parent.destroy()
        open_email_settings(parent.master, app_instance)
    tk.Button(win, text="保存", command=save, bg='#4CAF50', fg='white', relief=tk.FLAT).pack(pady=10)
    tk.Button(win, text="取消", command=win.destroy, bg='#3c3c3c', fg='white', relief=tk.FLAT).pack()

class App:
    def __init__(self, root):
        self.root = root
        root.title("多功能AI助手")
        root.geometry("1000x750")
        root.configure(bg='#1e1e1e')
        self.config = load_config()
        self.email_accounts = self.config.get("email_accounts", [])
        self.fetch_count = self.config.get("fetch_count", 5)
        self.pushplus_token = self.config.get("pushplus_token", "")
        self.current_email_index = tk.IntVar(value=0)
        # 标题
        title_label = tk.Label(root, text="⚡ 多功能AI助手", font=("微软雅黑", 18, "bold"), bg='#1e1e1e', fg='#ffffff')
        title_label.pack(pady=(20,10))
        # 邮箱选择区域
        self.create_email_selector()
        # 自定义分析需求（用于Excel客户分析）
        prompt_frame = tk.Frame(root, bg='#1e1e1e')
        prompt_frame.pack(pady=5, fill='x', padx=20)
        tk.Label(prompt_frame, text="Excel分析自定义需求（留空则使用默认）:", bg='#1e1e1e', fg='white', font=("微软雅黑", 10)).pack(anchor='w')
        self.user_prompt_entry = tk.Entry(prompt_frame, width=100, bg='#3c3c3c', fg='white', insertbackground='white')
        self.user_prompt_entry.pack(fill='x', pady=5)
        self.user_prompt_entry.insert(0, "筛选出可能需要使用选择性波峰焊设备的企业")
        # 产品分析输入框
        product_frame = tk.Frame(root, bg='#1e1e1e')
        product_frame.pack(pady=5, fill='x', padx=20)
        tk.Label(product_frame, text="产品描述（用于产品分析）:", bg='#1e1e1e', fg='white', font=("微软雅黑", 10)).pack(anchor='w')
        self.product_entry = tk.Entry(product_frame, width=100, bg='#3c3c3c', fg='white', insertbackground='white')
        self.product_entry.pack(fill='x', pady=5)
        # 按钮区域
        btn_frame = tk.Frame(root, bg='#1e1e1e')
        btn_frame.pack(pady=10)
        btn_style = {"bg": '#3c3c3c', "fg": "white", "font": ("微软雅黑", 10), "relief": tk.FLAT, "padx": 12, "pady": 5, "cursor": "hand2"}
        # 第一行按钮
        self.btn1 = tk.Button(btn_frame, text="📄 会议纪要", command=self.run_meeting, **btn_style)
        self.btn1.grid(row=0, column=0, padx=5)
        self.btn2 = tk.Button(btn_frame, text="✉️ 邮件提取", command=self.run_email, **btn_style)
        self.btn2.grid(row=0, column=1, padx=5)
        self.btn3 = tk.Button(btn_frame, text="📝 周报生成", command=self.run_weekly, **btn_style)
        self.btn3.grid(row=0, column=2, padx=5)
        self.btn4 = tk.Button(btn_frame, text="📊 Excel客户分析", command=self.run_excel_custom, **btn_style)
        self.btn4.grid(row=0, column=3, padx=5)
        self.btn5 = tk.Button(btn_frame, text="📦 产品分析", command=self.run_product, **btn_style)
        self.btn5.grid(row=0, column=4, padx=5)
        self.btn6 = tk.Button(btn_frame, text="📄 合同分析", command=self.run_contract, **btn_style)
        self.btn6.grid(row=0, column=5, padx=5)
        self.btn7 = tk.Button(btn_frame, text="🖼️ 图片处理", command=lambda: open_image_processor(self.root), **btn_style)
        self.btn7.grid(row=0, column=6, padx=5)
        self.btn8 = tk.Button(btn_frame, text="⚙️ 邮箱设置", command=lambda: open_email_settings(self.root, self), **btn_style)
        self.btn8.grid(row=0, column=7, padx=5)
        self.btn9 = tk.Button(btn_frame, text="📱 发到微信", command=self.send_last_to_wechat, **btn_style)
        self.btn9.grid(row=0, column=8, padx=5)
        # 输出文本框
        self.text_area = scrolledtext.ScrolledText(root, wrap=tk.WORD, font=("Consolas", 10), bg='#2d2d2d', fg='#d4d4d4', insertbackground='white')
        self.text_area.pack(padx=20, pady=10, fill=tk.BOTH, expand=True)
        # 状态栏
        self.status = tk.Label(root, text="✅ 就绪", bd=1, relief=tk.SUNKEN, anchor=tk.W, bg='#2d2d2d', fg='#cccccc')
        self.status.pack(side=tk.BOTTOM, fill=tk.X)
        self.last_result = ""

    def create_email_selector(self):
        if hasattr(self, 'selector_frame'):
            self.selector_frame.destroy()
        self.selector_frame = tk.Frame(self.root, bg='#1e1e1e')
        self.selector_frame.pack(pady=5)
        tk.Label(self.selector_frame, text="选择邮箱账号:", bg='#1e1e1e', fg='white', font=("微软雅黑", 10)).pack(side='left', padx=5)
        if self.email_accounts:
            self.account_var = tk.StringVar()
            self.account_combo = ttk.Combobox(self.selector_frame, textvariable=self.account_var, values=[acc["name"] for acc in self.email_accounts], state="readonly", width=30)
            self.account_combo.pack(side='left', padx=5)
            if self.email_accounts:
                self.account_combo.current(0)
        else:
            self.account_var = None
            tk.Label(self.selector_frame, text="⚠️ 未配置邮箱账号，请点击⚙️邮箱设置添加", bg='#1e1e1e', fg='orange').pack(side='left', padx=5)

    def refresh_accounts(self):
        self.config = load_config()
        self.email_accounts = self.config.get("email_accounts", [])
        self.fetch_count = self.config.get("fetch_count", 5)
        self.pushplus_token = self.config.get("pushplus_token", "")
        self.create_email_selector()

    def update_status(self, msg):
        self.status.config(text=msg)
        self.root.update_idletasks()

    def append_result(self, title, result):
        self.last_result = result
        self.text_area.insert(tk.END, f"\n{'='*60}\n【{title}】\n{'='*60}\n")
        self.text_area.insert(tk.END, result + "\n\n")
        self.text_area.see(tk.END)

    def run_task(self, func, title):
        def task():
            self.update_status("⏳ 处理中...")
            try:
                res = func()
                self.root.after(0, lambda: self.append_result(title, res))
                self.update_status("✅ 完成")
            except Exception as e:
                self.root.after(0, lambda: messagebox.showerror("错误", str(e)))
                self.update_status("❌ 失败")
        threading.Thread(target=task, daemon=True).start()

    def run_meeting(self):
        self.run_task(summarize_meeting, "会议纪要整理结果")

    def run_email(self):
        if not self.email_accounts:
            self.run_task(lambda: "未配置邮箱账号，请点击⚙️邮箱设置添加", "邮件提取结果")
            return
        idx = self.account_combo.current()
        if idx < 0:
            self.run_task(lambda: "请选择一个邮箱账号", "邮件提取结果")
            return
        selected_acc = self.email_accounts[idx]
        self.run_task(lambda: extract_email(selected_acc), "邮件重点提取结果")

    def run_weekly(self):
        self.run_task(generate_weekly, "周报生成结果")

    def run_excel_custom(self):
        file_path = filedialog.askopenfilename(title="选择Excel文件", filetypes=[("Excel文件", "*.xlsx *.xls")])
        if file_path:
            custom = self.user_prompt_entry.get()
            self.run_task(lambda: analyze_excel_custom(file_path, custom), "Excel客户分析结果")

    def run_product(self):
        desc = self.product_entry.get().strip()
        if not desc:
            messagebox.showwarning("提示", "请输入产品描述")
            return
        self.run_task(lambda: analyze_product(desc), "产品分析结果")

    def run_contract(self):
        file_path = filedialog.askopenfilename(title="选择合同文件", filetypes=[("合同文件", "*.docx *.xlsx *.xls")])
        if file_path:
            self.run_task(lambda: analyze_contract(file_path), "合同分析结果")

    def send_last_to_wechat(self):
        if not self.last_result:
            messagebox.showinfo("提示", "没有可发送的结果，请先执行一个功能")
            return
        token = self.pushplus_token
        if not token:
            token = load_config().get("pushplus_token", "")
            if not token:
                messagebox.showwarning("提示", "未配置 pushplus token，请点击⚙️邮箱设置填写")
                return
        def send():
            msg = send_to_wechat(self.last_result, token)
            self.root.after(0, lambda: messagebox.showinfo("发送结果", msg))
        threading.Thread(target=send, daemon=True).start()

if __name__ == "__main__":
    root = tk.Tk()
    app = App(root)
    root.mainloop()